from openpyxl import load_workbook
import openpyxl

import os
import re

def collect_data():
    
    all_recipes = []
    pattern = re.compile(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)')

    # Load files
    for file in os.listdir("."):

        # Ignore non-xlsx files
        if not file.endswith(".xlsx"):
            continue
        
        # Load file
        print(file)
        wb = load_workbook(file)
        ws = wb.active

        # Get all the values
        all_recipes.extend([
            pattern.match(cell.value).groups()
            for row in ws.iter_rows(values_only=False)
            for cell in row
            if "HYPERLINK" in cell.value
        ])

    return all_recipes


def tabulate_data(all_recipes):

    # Sort in order
    all_recipes.sort(key=lambda x: x[1])

    return all_recipes

def store_data():
    ...

# Load all the files
all_recipes = collect_data()

# Tabulated data
tabulated_data = tabulate_data(all_recipes)

store_data()